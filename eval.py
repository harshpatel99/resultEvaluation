import PyPDF2
from io import BytesIO
import pandas as pd
import operator
import openpyxl

from openpyxl.styles.borders import Border,Side
from openpyxl.styles import Alignment

thinBorder = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))

totalSubjects = 10
headerRow = 2
summaryRow = 0

pdf = open('SE IT 2015.pdf','rb')
pdfReader = PyPDF2.PdfFileReader(pdf)

total = pdfReader.numPages

pageOne = pdfReader.getPage(1).extractText()
details = {}

flag = 0
for i in range(0,total):
    page = pdfReader.getPage(i).extractText()
    j=1
    index = 393
    while j < 3:
        if page[index+12] == '.':
            j+=1
            flag = 1
            break
        examSeatNo = page[index:index+12].strip()
        details[examSeatNo] = {}
        details[examSeatNo]['examSeatNo'] = page[index:index+12].strip()
        index+=12
        details[examSeatNo]['name'] = page[index:index+40].strip()
        index+=40
        val=25
        while not(page[index+val].isdigit()):
            val+=1
        details[examSeatNo]['motherName'] = page[index:index + val].strip()
        index+=val
        #index+=25
        #if page[index].isdigit():
        details[examSeatNo]['prn'] = page[index:index+10]
        #else:
            #tempIndex = page.find(" ",index) - 9
            #details[examSeatNo]['prn'] = page[tempIndex:page.find(" ",index)]
            #index+=2
        index+=292
        totalSem = 2
        if page[0:page.find("SEM.:2")] == -1:
            totalSem = 1
        k = 1
        while k <= totalSem:
            details[examSeatNo]['sem{}'.format(page[index:page.find(" ",index)])] = {}
            semNo = 'sem{}'.format(page[index:page.find(" ",index)])
            index+=5
            for m in range(1,11):
                if not page[index].isdigit():
                    break
                subjectID = ""
                if page[index-1:index] == ' ':
                    subjectID = page[index:page.find(" ",index)].strip()
                else:
                    index += 85
                    if m == 10:
                        index += 9
                    else:
                        index +=5
                    continue
                    #subjectID = page[index-1:page.find(" ",index-1)]
                #subjectID = page[index:page.find(" ",index)]
                if subjectID not in details[examSeatNo][semNo].keys():
                    details[examSeatNo][semNo][subjectID] = {}
                else:
                    subjectID = '99' + subjectID
                    details[examSeatNo][semNo][subjectID] = {}
                details[examSeatNo][semNo][subjectID]['subjectId'] = subjectID
                index+=10

                details[examSeatNo][semNo][subjectID]['oe'] = page[index:index+3].strip()
                index+=9
                details[examSeatNo][semNo][subjectID]['th'] = page[index:index+3].strip()
                index+=18
                details[examSeatNo][semNo][subjectID]['tw'] = page[index:index+3].strip()
                index+=9
                if details[examSeatNo][semNo][subjectID]['tw'] != '---' and not details[examSeatNo][semNo][subjectID]['tw'].isdigit():
                    details[examSeatNo][semNo][subjectID]['pr'] = page[index:index+4].strip()
                else:
                    details[examSeatNo][semNo][subjectID]['pr'] = page[index:index+3].strip()
                index+=9
                details[examSeatNo][semNo][subjectID]['oral'] = page[index:index+3].strip()
                index+=9
                details[examSeatNo][semNo][subjectID]['tt'] = page[index:index+2].strip()
                index+=5
                details[examSeatNo][semNo][subjectID]['crd'] = page[index:index+2].strip()
                index+=5
                details[examSeatNo][semNo][subjectID]['gr'] = page[index:index+2].strip()
                index+=5
                details[examSeatNo][semNo][subjectID]['gradePoints'] = page[index:index+2].strip()
                index+=6
                details[examSeatNo][semNo][subjectID]['cr'] = page[index:index+2].strip()
                if m == 10:
                    index+=9
                else:
                    index+=5

            #index+=9
            k+=1
        #print(details)
        #print(pd.DataFrame(details))
        #index+=16
        #if page[index-1] == '-':
        #print(index)
        details[examSeatNo]['pointer'] = page[page.find(":",index)+2:page.find(',',index)]
        index+=16
        if page[page.find(':',index) + 4] == 'R':
            details[examSeatNo]['totalCredits'] = page[page.find(':',index)+2:page.find('RESERVED',index)-7]
        elif page.find(':',index) + 4 == '.':
            details[examSeatNo]['totalCredits'] = page[page.find(':',index)+2:page.find('.',index)]
        else:
            details[examSeatNo]['totalCredits'] = page[page.find(':',index)+2:page.find('  ',index)]
        #else:
        #    details[examSeatNo]['pointer'] = page[index-4:index]

        index = page.find('DISTRIBUTION...........') + 24
        #index+=182


        j+=1
        #index+=2024
    if flag == 1:
        break

#print(details)
#print(pd.DataFrame(details))
#df = pd.DataFrame(data=details, index=[0])
#df.to_excel('dict1.xlsx')

def getValueWithDataType(val):
    if(val.isdigit()):
        return int(val)
    else:
        return val

for key in details.keys():
    if details[key]['pointer'] != '--':
        details[key]['pointer'] = float(details[key]['pointer'])
    else:
        details[key]['pointer'] = float(0)

sum = 0
sem = 'sem2'
for key in details.keys():
    sum = 0
    if 'sem2' in details[key]:
        sem = 'sem2'
    else:
        sem = 'sem1'
    for subKey in details[key][sem].keys():
        #print(sum)
        if details[key][sem][subKey]['oe'].isdigit():
            sum += getValueWithDataType(details[key][sem][subKey]['oe'])
        if details[key][sem][subKey]['th'].isdigit():
            sum += getValueWithDataType(details[key][sem][subKey]['th'])
        if details[key][sem][subKey]['tw'].isdigit():
            sum += getValueWithDataType(details[key][sem][subKey]['tw'])
        if details[key][sem][subKey]['pr'].isdigit():
            sum += getValueWithDataType(details[key][sem][subKey]['pr'])
    details[key]['total'] = sum
    #print(sum)
    #print('\n')


#Keymax = max(val, key=val.get)
#print(Keymax)

#print(details[max(details, key=lambda x: details[x]['pointer'])])



workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = 'Analysis Regular'

def colNumString(n):
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string

prevCol = 4
superCol = 4
row = 3
for key,values in details.items():
    #sheet.cell(row=row, column=1, value=key)
    column = 2
    for element in values:
        if element == 'motherName' or element == 'prn':
            continue
        if element == 'sem1' or element == 'sem2':
            for subject in values[element]:
                #sheet.cell(row=2, column=column, value=subject)
                #sheet.merge_cells('{}1:{}1'.format(colNumString(column),colNumString(column+5)))
                #sheet.merge_cells('D1:H1')

                for subKeys in values[element][subject]:

                    sheet.column_dimensions[colNumString(column)].width = 5

                    if subKeys == 'subjectId':
                        sheet.cell(row=1, column=column, value=getValueWithDataType(values[element][subject][subKeys]))
                        prevCol = column
                        #sheet.merge_cells('D1:H1')
                        continue
                    elif subKeys == 'crd' or subKeys == 'gradePoints':
                        continue
                    sheet.cell(row=headerRow, column=column, value=subKeys.upper())

                    #sheet.cell(row=1, column=superCol, value=subKeys)

                    if values[element][subject][subKeys] != '---':
                        val = getValueWithDataType(values[element][subject][subKeys])
                        sheet.cell(row=row, column=column, value=val).border = thinBorder
                        sheet.cell(row=row, column=column).alignment = Alignment(horizontal='center')
                        column += 1
                #print(subject)
                #sheet.cell(row=row, column=column, value=values[element][subject])
                #column += 1
            continue
        else:
            if element == 'examSeatNo':
                sheet.column_dimensions[colNumString(column)].width = 12
            else:
                sheet.column_dimensions[colNumString(column)].width = 40
            if element == 'pointer':
                sheet.column_dimensions[colNumString(column)].width = 20
                sheet.cell(row=headerRow, column=column, value='Pointer').border = thinBorder
                sheet.cell(row=headerRow, column=column).alignment = Alignment(horizontal='center')
                sheet.cell(row=row, column=column, value=values[element]).border = thinBorder
                sheet.cell(row=row, column=column).alignment = Alignment(horizontal='center')
                column += 1
                sheet.cell(row=headerRow, column=column, value='Class').border = thinBorder
                sheet.cell(row=headerRow, column=column).alignment = Alignment(horizontal='center')
                sheet.cell(row=row, column=column, value='=IF({}{}=0,"Fail",IF({}{}>7.74,"Dist",IF({}{}>6.74,"FC",IF({}{}>6.24,"HSC",IF({}{}>5.4,"SC","Pass")))))'.format(colNumString(column-1),row,colNumString(column-1),row,colNumString(column-1),row,colNumString(column-1),row,colNumString(column-1),row)).border = thinBorder
                sheet.cell(row=row, column=column).alignment = Alignment(horizontal='center')
            else:
                sheet.cell(row=row, column=column, value=values[element]).border = thinBorder
                sheet.cell(row=row, column=column).alignment = Alignment(horizontal='center')
            column += 1
    row += 1

"""sheet['{}{}'.format(colNumString(colFormula),rowFormula)] = '=COUNTIF(G$9:G$78,"O")'
sheet['D95'] = '=COUNTIF(G$9:G$78,"O")'
sheet['D95'] = '=COUNTIF(G$9:G$78,"O")'
sheet['D95'] = '=COUNTIF(G$9:G$78,"O")'
sheet['D95'] = '=COUNTIF(G$9:G$78,"O")'"""


startCol = 3
colFormula = 4
rowFormula = row+5
gradeCol = 7
summaryRow = row+5

r = row+5
sheet['{}{}'.format(colNumString(column-3),r)] = 'DISTICTION'
sheet['{}{}'.format(colNumString(column-2),r)] = '=COUNTIF({}${}:{}${},"Dist")'.format(colNumString(column-2),headerRow+1,colNumString(column-2),row)
r += 1
sheet['{}{}'.format(colNumString(column-3),r)] = 'FIRST CLASS'
sheet['{}{}'.format(colNumString(column-2),r)] = '=COUNTIF({}${}:{}${},"FC")'.format(colNumString(column-2),headerRow+1,colNumString(column-2),row)
r += 1
sheet['{}{}'.format(colNumString(column-3),r)] = 'HIGH. SEC.'
sheet['{}{}'.format(colNumString(column-2),r)] = '=COUNTIF({}${}:{}${},"HSC")'.format(colNumString(column-2),headerRow+1,colNumString(column-2),row)
r += 1
sheet['{}{}'.format(colNumString(column-3),r)] = 'SEC.'
sheet['{}{}'.format(colNumString(column-2),r)] = '=COUNTIF({}${}:{}${},"SC")'.format(colNumString(column-2),headerRow+1,colNumString(column-2),row)
r += 1
sheet['{}{}'.format(colNumString(column-3),r)] = 'PASS CLASS'
sheet['{}{}'.format(colNumString(column-2),r)] = '=COUNTIF({}${}:{}${},"Pass")'.format(colNumString(column-2),headerRow+1,colNumString(column-2),row)
r += 1
sheet['{}{}'.format(colNumString(column-3),r)] = 'FAIL'
sheet['{}{}'.format(colNumString(column-2),r)] = '=COUNTIF({}${}:{}${},"Fail")'.format(colNumString(column-2),headerRow+1,colNumString(column-2),row)
r += 1
sheet['{}{}'.format(colNumString(column-3),r)] = 'TOTAL PASS'
sheet['{}{}'.format(colNumString(column-2),r)] = '=SUM({}{}:{}{})'.format(colNumString(column-2),summaryRow,colNumString(column-2),r-2)
r += 1
sheet['{}{}'.format(colNumString(column-3),r)] = 'TOTAL APPEARED'
sheet['{}{}'.format(colNumString(column-2),r)] = '=SUM({}{}:{}{})'.format(colNumString(column-2),summaryRow,colNumString(column-2),r-2)
r += 1
sheet['{}{}'.format(colNumString(column-3),r)] = 'CLEAR STUDENT %'
sheet['{}{}'.format(colNumString(column-2),r)] = '=({}{}/{}{})*100'.format(colNumString(column-2),r-2,colNumString(column-2),r-1)
r += 1

sheet['{}{}'.format(colNumString(colFormula-1),rowFormula)] = 'Grade O'
rowFormula += 1
sheet['{}{}'.format(colNumString(colFormula-1),rowFormula)] = 'Grade A+'
rowFormula += 1
sheet['{}{}'.format(colNumString(colFormula-1),rowFormula)] = 'Grade A'
rowFormula += 1
sheet['{}{}'.format(colNumString(colFormula-1),rowFormula)] = 'Grade B+'
rowFormula += 1
sheet['{}{}'.format(colNumString(colFormula-1),rowFormula)] = 'Grade B'
rowFormula += 1
sheet['{}{}'.format(colNumString(colFormula-1),rowFormula)] = 'Grade C'
rowFormula += 1
sheet['{}{}'.format(colNumString(colFormula-1),rowFormula)] = 'Grade P'
rowFormula += 1
sheet['{}{}'.format(colNumString(colFormula-1),rowFormula)] = 'Grade F'
rowFormula += 1
sheet['{}{}'.format(colNumString(colFormula-1),rowFormula)] = 'Total Pass'
rowFormula += 1
sheet['{}{}'.format(colNumString(colFormula-1),rowFormula)] = 'Total Fail'
rowFormula += 1
sheet['{}{}'.format(colNumString(colFormula-1),rowFormula)] = '% Pass'
rowFormula += 1
sheet['{}{}'.format(colNumString(colFormula-1),rowFormula)] = '% Fail'

for subKeys in range(0,totalSubjects):
    rowFormula = row+5
    sheet['{}{}'.format(colNumString(colFormula),rowFormula)] = '=COUNTIF({}${}:{}${},"O")'.format(colNumString(gradeCol),startCol,colNumString(gradeCol),len(details)+2)
    rowFormula += 1
    sheet['{}{}'.format(colNumString(colFormula),rowFormula)] = '=COUNTIF({}${}:{}${},"A+")'.format(colNumString(gradeCol),startCol,colNumString(gradeCol),len(details)+2)
    rowFormula += 1
    sheet['{}{}'.format(colNumString(colFormula),rowFormula)] = '=COUNTIF({}${}:{}${},"A")'.format(colNumString(gradeCol),startCol,colNumString(gradeCol),len(details)+2)
    rowFormula += 1
    sheet['{}{}'.format(colNumString(colFormula),rowFormula)] = '=COUNTIF({}${}:{}${},"B+")'.format(colNumString(gradeCol),startCol,colNumString(gradeCol),len(details)+2)
    rowFormula += 1
    sheet['{}{}'.format(colNumString(colFormula),rowFormula)] = '=COUNTIF({}${}:{}${},"B")'.format(colNumString(gradeCol),startCol,colNumString(gradeCol),len(details)+2)
    rowFormula += 1
    sheet['{}{}'.format(colNumString(colFormula),rowFormula)] = '=COUNTIF({}${}:{}${},"C")'.format(colNumString(gradeCol),startCol,colNumString(gradeCol),len(details)+2)
    rowFormula += 1
    sheet['{}{}'.format(colNumString(colFormula),rowFormula)] = '=COUNTIF({}${}:{}${},"P")'.format(colNumString(gradeCol),startCol,colNumString(gradeCol),len(details)+2)
    rowFormula += 1
    sheet['{}{}'.format(colNumString(colFormula),rowFormula)] = '=COUNTIF({}${}:{}${},"F")'.format(colNumString(gradeCol),startCol,colNumString(gradeCol),len(details)+2)
    rowFormula += 1
    sheet['{}{}'.format(colNumString(colFormula),rowFormula)] = '=SUM({}{}:{}{})'.format(colNumString(colFormula),row+5,colNumString(colFormula),row+12)
    rowFormula += 1
    sheet['{}{}'.format(colNumString(colFormula),rowFormula)] = '={}{}'.format(colNumString(colFormula),rowFormula-2)
    rowFormula += 1
    sheet['{}{}'.format(colNumString(colFormula),rowFormula)] = '={}{}/({}{}+{}{})%'.format(colNumString(colFormula),rowFormula-1,colNumString(colFormula),rowFormula-1,colNumString(colFormula),rowFormula-2)
    rowFormula += 1
    sheet['{}{}'.format(colNumString(colFormula),rowFormula)] = '={}{}/({}{}+{}{})%'.format(colNumString(colFormula),rowFormula-3,colNumString(colFormula),rowFormula-3,colNumString(colFormula),rowFormula-2)
    rowFormula += 1
    colFormula += 5
    gradeCol += 5

#Overall sheet
sheet = workbook.create_sheet(title="Overall")

from heapq import nlargest
topTen = nlargest(10, details, key = lambda x: details[x]['pointer'])

r = 30
c = 10
rank = 1

sheet.column_dimensions[colNumString(c)].width = 8
sheet.cell(row=r-1, column=c, value='Rank').border = thinBorder
sheet.cell(row=r-1, column=c).alignment = Alignment(horizontal='center')
sheet.column_dimensions[colNumString(c+1)].width = 35
sheet.cell(row=r-1, column=c+1, value='Name of Student').border = thinBorder
sheet.cell(row=r-1, column=c+1).alignment = Alignment(horizontal='center')
sheet.column_dimensions[colNumString(c+2)].width = 8
sheet.cell(row=r-1, column=c+2, value='SGPA').border = thinBorder
sheet.cell(row=r-1, column=c+2).alignment = Alignment(horizontal='center')
sheet.column_dimensions[colNumString(c+3)].width = 14
sheet.cell(row=r-1, column=c+3, value='Class').border = thinBorder
sheet.cell(row=r-1, column=c+3).alignment = Alignment(horizontal='center')

for i in topTen:
    c = 10
    pointer = details[i]['pointer']

    sheet.cell(row=r, column=c, value=rank).border = thinBorder
    sheet.cell(row=r, column=c).alignment = Alignment(horizontal='center')
    c += 1
    sheet.cell(row=r, column=c, value=details[i]['name']).border = thinBorder
    c += 1
    sheet.cell(row=r, column=c, value=pointer).border = thinBorder
    sheet.cell(row=r, column=c).alignment = Alignment(horizontal='center')
    c += 1
    sClass = 'Distinction'
    if pointer > 7.74:
        sClass = 'Distinction'
    elif pointer > 6.74:
        sClass = 'First Class'
    elif pointer > 6.24:
        sClass = 'High Sec'
    elif pointer > 5.4:
        sClass = 'Second Class'
    else:
        sClass = 'Pass'
    sheet.cell(row=r, column=c, value=sClass).border = thinBorder
    sheet.cell(row=r, column=c).alignment = Alignment(horizontal='center')
    rank += 1
    r += 1

#Percentage Sheet
row = 2
sheet = workbook.create_sheet(title="Percentage")
for key,values in details.items():
    column = 1
    for element in values:
        sheet.column_dimensions[colNumString(column)].width = 20
        if element == 'examSeatNo':
            sheet.cell(row=headerRow, column=column, value='Exam Seat No.').border = thinBorder
            sheet.cell(row=headerRow, column=column).alignment = Alignment(horizontal='center')
            sheet.cell(row=row, column=column, value=values[element]).border = thinBorder
            sheet.cell(row=row, column=column).alignment = Alignment(horizontal='center')
            column += 1
        elif element == 'name':
            sheet.cell(row=headerRow, column=column, value='Name').border = thinBorder
            sheet.cell(row=headerRow, column=column).alignment = Alignment(horizontal='center')
            sheet.cell(row=row, column=column, value=values[element]).border = thinBorder
            column += 1
        elif element == 'total':
            sheet.column_dimensions[colNumString(column)].width = 15
            sheet.cell(row=headerRow, column=column, value='Total Marks').border = thinBorder
            sheet.cell(row=headerRow, column=column).alignment = Alignment(horizontal='center')
            sheet.cell(row=row, column=column, value=values[element]).border = thinBorder
            sheet.cell(row=row, column=column).alignment = Alignment(horizontal='center')
            column += 1
    sheet.column_dimensions[colNumString(column)].width = 15
    sheet.cell(row=headerRow, column=column, value='Total Marks').border = thinBorder
    sheet.cell(row=headerRow, column=column).alignment = Alignment(horizontal='center')
    sheet.cell(row=row, column=column, value='=ROUND({}{}/750*100,2)'.format(colNumString(column-1),row)).border = thinBorder
    sheet.cell(row=row, column=column).alignment = Alignment(horizontal='center')
    column += 1
    row += 1

workbook.save(filename="my_workbook.xlsx")


"""row = 2
col = 1
sheet.cell(row=2, column=col, value='Sr. No.')
col+=1
sheet.cell(row=2, column=col, value='Exam Seat No.')
col+=1
sheet.cell(row=2, column=col, value='Name of Student')
col+=1
sheet.cell(row=2, column=col, value='PRN')
col+=1
sheet.cell(row=2, column=col, value='Total Cerdits (TH)')
col+=1
sheet.cell(row=2, column=col, value='Total Cerdits (PR)')
col+=1
sheet.cell(row=2, column=col, value='Class')
col+=1
sheet.cell(row=2, column=col, value='No. of Backlogs (TH)')
col+=1
sheet.cell(row=2, column=col, value='No. of Backlogs (PR)')
col+=1
sheet.cell(row=2, column=col, value='Total Backlogs')
col+=1
sheet.cell(row=2, column=col, value='Pointer')"""

#767
#print(index)
#print(pageOne.find("DISTRIBUTION..........."))
#print(pdfReader.getPage(10).extractText())
